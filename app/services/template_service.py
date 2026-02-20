"""
Template service — Excel plantilla (template) generator for INEI data formats.

Generates downloadable ``.xlsx`` template files for each of the 10 standardised
INEI import formats using ``openpyxl``.  Each workbook contains:

* **Data sheet** — a context header block (UE, Meta, Año), styled column headers
  (bold, INEI blue ``#3b82f6``, white text, borders), and empty data rows.
* **Instrucciones sheet** — plain-text filling instructions for end users.

Public API
----------
- ``FORMATO_CATALOG`` — list of format metadata dicts (key, nombre, etc.).
- ``get_formato_catalog()`` — returns the catalog (safe copy).
- ``generate_template(formato_key, output_path)`` — writes a single ``.xlsx`` file.
- ``generate_all_templates(plantillas_dir)`` — writes all 10 templates to a directory.

Design notes
------------
- Uses ``openpyxl`` directly (not ``xlsxwriter`` or ``pandas``) to keep the
  API simple and avoid a secondary dependency path.
- Column header row is placed at ``fila_inicio - 1`` (1-based), matching the
  real file layout that parsers expect.
- Context rows (rows 1-4) are always written at fixed positions regardless of
  ``fila_inicio``, with blank filler rows in between when needed.
- The service is intentionally side-effect free — it writes only to the path
  given and returns it; no DB access, no FastAPI state.
"""

from __future__ import annotations

import copy
import logging
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Design tokens (kept in sync with CLAUDE.md and excel_exporter.py)
# ---------------------------------------------------------------------------

_HEX_PRIMARY = "3b82f6"        # INEI blue — column header background
_HEX_WHITE = "FFFFFF"          # white text on blue headers
_HEX_LABEL_BG = "EFF6FF"       # very light blue for context label cells
_HEX_LABEL_TEXT = "1E3A5F"     # dark navy for context label text
_HEX_TITLE_BG = "1E3A5F"       # dark navy for format title row
_HEX_BORDER = "CBD5E1"         # slate-300 for all borders

_CURRENT_YEAR: int = date.today().year


# ---------------------------------------------------------------------------
# Format catalog
# ---------------------------------------------------------------------------

#: Complete catalog of the 10 INEI Excel import formats.
#: Each entry is a plain dict with the following keys:
#:
#: * ``key``         — unique identifier used in API calls and file names.
#: * ``nombre``      — human-readable format name (Spanish).
#: * ``descripcion`` — one-line summary of the format's purpose.
#: * ``hoja``        — sheet name as it appears in real source files.
#: * ``columnas``    — ordered list of column header strings.
#: * ``fila_inicio`` — 1-based row number where data rows begin (mirrors parsers).
FORMATO_CATALOG: list[dict[str, Any]] = [
    {
        "key": "cuadro_ao_meta",
        "nombre": "Cuadro AO-META",
        "descripcion": "Datos maestros: 85 actividades operativas y sus relaciones con metas.",
        "hoja": "Cuadro AO-Meta",
        "columnas": [
            "N°",
            "Codigo CEPLAN",
            "Nombre AO",
            "Codigo Meta",
            "Descripcion Meta",
            "Area Responsable",
        ],
        "fila_inicio": 7,
    },
    {
        "key": "tablas",
        "nombre": "Tablas de Referencia",
        "descripcion": "Datos maestros: 569 clasificadores de gasto y tipos de referencia.",
        "hoja": "Tablas",
        "columnas": [
            "Clasificador",
            "Tipo Generico",
            "Tipo Especifico",
            "Sub Tipo",
            "Descripcion",
            "Estado",
        ],
        "fila_inicio": 5,
    },
    {
        "key": "formato1",
        "nombre": "Formato 1 - Programacion Presupuestal",
        "descripcion": "Programacion presupuestal anual por clasificador de gasto (23 columnas desde F8).",
        "hoja": "Formato 1",
        "columnas": [
            "Clasificador",
            "Descripcion",
            "PIA",
            "PIM",
            "Ene",
            "Feb",
            "Mar",
            "Abr",
            "May",
            "Jun",
            "Jul",
            "Ago",
            "Sep",
            "Oct",
            "Nov",
            "Dic",
            "Total",
        ],
        "fila_inicio": 8,
    },
    {
        "key": "formato2",
        "nombre": "Formato 2 - Programacion por Tareas",
        "descripcion": "Programacion a nivel de tarea (19 columnas desde F8).",
        "hoja": "Formato 2",
        "columnas": [
            "Cod Meta",
            "Desc Meta",
            "Cod AO",
            "Desc AO",
            "Cod Tarea",
            "Desc Tarea",
            "Clasificador",
            "Desc Clasificador",
            "PIM",
            "Ene",
            "Feb",
            "Mar",
            "Abr",
            "May",
            "Jun",
            "Jul",
            "Ago",
            "Sep",
            "Oct",
            "Nov",
            "Dic",
        ],
        "fila_inicio": 8,
    },
    {
        "key": "formato3",
        "nombre": "Formato 3 - Tareas con Justificacion",
        "descripcion": "Tareas con campos de justificacion y observaciones.",
        "hoja": "Formato 3",
        "columnas": [
            "Cod Meta",
            "Desc Meta",
            "Cod AO",
            "Desc AO",
            "Cod Tarea",
            "Desc Tarea",
            "Clasificador",
            "Desc Clasificador",
            "PIM",
            "Programado",
            "Ejecutado",
            "Saldo",
            "% Avance",
            "Justificacion",
            "Observaciones",
        ],
        "fila_inicio": 8,
    },
    {
        "key": "formato04",
        "nombre": "Formato 04 - Modificaciones Presupuestales",
        "descripcion": "Registro de modificaciones presupuestales (6 columnas desde F8).",
        "hoja": "Formato 04",
        "columnas": [
            "Clasificador",
            "Descripcion",
            "Asignado",
            "Habilitadora",
            "Habilitada",
            "PIM Resultante",
        ],
        "fila_inicio": 8,
    },
    {
        "key": "formato5a",
        "nombre": "Formato 5.A - Programacion AO",
        "descripcion": "Programacion mensual de actividades operativas (22 columnas desde F12).",
        "hoja": "Formato 5.A",
        "columnas": [
            "Codigo AO",
            "Nombre AO",
            "Ene",
            "Feb",
            "Mar",
            "Abr",
            "May",
            "Jun",
            "Jul",
            "Ago",
            "Sep",
            "Oct",
            "Nov",
            "Dic",
            "Total Programado",
        ],
        "fila_inicio": 12,
    },
    {
        "key": "formato5b",
        "nombre": "Formato 5.B - Ejecucion AO",
        "descripcion": "Ejecucion mensual triple (programado/ejecutado/saldo x 12 meses, 45 columnas desde F12).",
        "hoja": "Formato 5.B",
        "columnas": [
            "Codigo AO",
            "Nombre AO",
            # Programado (12 months)
            "Prog Ene", "Prog Feb", "Prog Mar", "Prog Abr",
            "Prog May", "Prog Jun", "Prog Jul", "Prog Ago",
            "Prog Sep", "Prog Oct", "Prog Nov", "Prog Dic",
            # Ejecutado (12 months)
            "Ejec Ene", "Ejec Feb", "Ejec Mar", "Ejec Abr",
            "Ejec May", "Ejec Jun", "Ejec Jul", "Ejec Ago",
            "Ejec Sep", "Ejec Oct", "Ejec Nov", "Ejec Dic",
            # Saldo (12 months)
            "Saldo Ene", "Saldo Feb", "Saldo Mar", "Saldo Abr",
            "Saldo May", "Saldo Jun", "Saldo Jul", "Saldo Ago",
            "Saldo Sep", "Saldo Oct", "Saldo Nov", "Saldo Dic",
            # Totals and summary
            "Total Prog",
            "Total Ejec",
            "Total Saldo",
            "PIM",
            "% Avance",
        ],
        "fila_inicio": 12,
    },
    {
        "key": "formato5_resumen",
        "nombre": "Formato 5 Resumen - Resumen Ejecucion",
        "descripcion": "Resumen de ejecucion por AO (20 columnas desde F7).",
        "hoja": "Formato 5 Resumen",
        "columnas": [
            "Codigo AO",
            "Nombre AO",
            "PIM",
            "CCP",
            "Compromiso Anual",
            "Devengado",
            "Girado",
            "Saldo",
            "% Avance PIM",
            "% Avance CCP",
            "Semaforo",
            "Ene",
            "Feb",
            "Mar",
            "Abr",
            "May",
            "Jun",
            "Jul",
            "Ago",
            "Sep",
            "Oct",
            "Nov",
            "Dic",
        ],
        "fila_inicio": 7,
    },
    {
        "key": "anexo01",
        "nombre": "Anexo 01 - Datos RRHH",
        "descripcion": "Datos de recursos humanos por unidad ejecutora.",
        "hoja": "Anexo 01",
        "columnas": [
            "N°",
            "DNI",
            "Apellidos y Nombres",
            "Cargo",
            "Area",
            "Regimen Laboral",
            "Tipo Contrato",
            "Fecha Inicio",
            "Fecha Fin",
            "Remuneracion Mensual",
            "Observaciones",
            "Estado",
        ],
        "fila_inicio": 8,
    },
]

# Fast lookup by key
_CATALOG_BY_KEY: dict[str, dict[str, Any]] = {
    fmt["key"]: fmt for fmt in FORMATO_CATALOG
}


# ---------------------------------------------------------------------------
# openpyxl style helpers
# ---------------------------------------------------------------------------

def _make_thin_border() -> Border:
    """Return a thin border on all four sides using the project slate colour.

    Returns:
        An ``openpyxl`` ``Border`` object with thin sides.
    """
    thin_side = Side(style="thin", color=_HEX_BORDER)
    return Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )


def _make_medium_border() -> Border:
    """Return a medium outer border for the column-header row.

    Returns:
        An ``openpyxl`` ``Border`` with medium sides.
    """
    medium_side = Side(style="medium", color=_HEX_BORDER)
    return Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=medium_side,
    )


def _apply_col_header_style(cell: Any) -> None:
    """Apply the INEI column-header style to a single openpyxl cell.

    Style: bold, white text, INEI blue (#3b82f6) background, all-sides thin
    border, centred horizontally and vertically, text-wrap enabled.

    Args:
        cell: An ``openpyxl`` ``Cell`` object to style in-place.
    """
    cell.font = Font(
        bold=True,
        color=_HEX_WHITE,
        size=10,
        name="Calibri",
    )
    cell.fill = PatternFill(
        fill_type="solid",
        fgColor=_HEX_PRIMARY,
    )
    cell.border = _make_thin_border()
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )


def _apply_context_label_style(cell: Any) -> None:
    """Apply the style for a context-area label cell (e.g. 'Unidad Ejecutora:').

    Style: bold, dark-navy text, very-light-blue background, all-sides thin
    border, right-aligned.

    Args:
        cell: An ``openpyxl`` ``Cell`` object to style in-place.
    """
    cell.font = Font(
        bold=True,
        color=_HEX_LABEL_TEXT,
        size=10,
        name="Calibri",
    )
    cell.fill = PatternFill(
        fill_type="solid",
        fgColor=_HEX_LABEL_BG,
    )
    cell.border = _make_thin_border()
    cell.alignment = Alignment(horizontal="right", vertical="center")


def _apply_context_value_style(cell: Any) -> None:
    """Apply the style for a context-area value cell (user fills this in).

    Style: regular weight, dark text, white background, all-sides thin border,
    left-aligned — clearly indicating an editable input cell.

    Args:
        cell: An ``openpyxl`` ``Cell`` object to style in-place.
    """
    cell.font = Font(color="111827", size=10, name="Calibri")
    cell.fill = PatternFill(fill_type="solid", fgColor=_HEX_WHITE)
    cell.border = _make_thin_border()
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_title_style(cell: Any) -> None:
    """Apply the format-title style (row 1 merged cell).

    Style: bold, white text, dark-navy background, font size 14, centred.

    Args:
        cell: An ``openpyxl`` ``Cell`` object to style in-place.
    """
    cell.font = Font(
        bold=True,
        color=_HEX_WHITE,
        size=14,
        name="Calibri",
    )
    cell.fill = PatternFill(fill_type="solid", fgColor=_HEX_TITLE_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Context header block writer
# ---------------------------------------------------------------------------

def _write_context_header(
    ws: Any,
    formato_nombre: str,
    fila_inicio: int,
    num_cols: int,
) -> None:
    """Write the four-row context header block to the given worksheet.

    Layout (all 1-based row indices):
    - Row 1  : Format title, merged across ``min(num_cols, 6)`` columns.
    - Row 2  : "Unidad Ejecutora:" label (col A) + empty value cell (col B).
    - Row 3  : "Meta Presupuestal:" label (col A) + empty value cell (col B).
    - Row 4  : "Año Fiscal:" label (col A) + current year as example (col B).
    - Rows 5 … (fila_inicio - 2): blank filler rows (no styling).
    - Row (fila_inicio - 1): Column header row — written by the caller.

    Args:
        ws: An ``openpyxl`` ``Worksheet`` object.
        formato_nombre: Human-readable format name for the title cell.
        fila_inicio: 1-based first data row (determines where headers go).
        num_cols: Total number of data columns (used for merge width).
    """
    # The title merge spans at most 6 columns so it stays compact on narrow sheets
    merge_end_col = min(num_cols, 6)
    merge_end_letter = get_column_letter(merge_end_col)

    # --- Row 1: title ---
    title_cell = ws["A1"]
    title_cell.value = formato_nombre
    _apply_title_style(title_cell)
    ws.row_dimensions[1].height = 28

    if merge_end_col > 1:
        ws.merge_cells(f"A1:{merge_end_letter}1")

    # --- Row 2: Unidad Ejecutora ---
    ws.row_dimensions[2].height = 18
    label_ue = ws["A2"]
    label_ue.value = "Unidad Ejecutora:"
    _apply_context_label_style(label_ue)
    value_ue = ws["B2"]
    value_ue.value = ""
    _apply_context_value_style(value_ue)

    # --- Row 3: Meta Presupuestal ---
    ws.row_dimensions[3].height = 18
    label_meta = ws["A3"]
    label_meta.value = "Meta Presupuestal:"
    _apply_context_label_style(label_meta)
    value_meta = ws["B3"]
    value_meta.value = ""
    _apply_context_value_style(value_meta)

    # --- Row 4: Año Fiscal ---
    ws.row_dimensions[4].height = 18
    label_anio = ws["A4"]
    label_anio.value = "Año Fiscal:"
    _apply_context_label_style(label_anio)
    value_anio = ws["B4"]
    value_anio.value = _CURRENT_YEAR
    _apply_context_value_style(value_anio)

    # --- Rows 5 … (fila_inicio - 2): blank filler ---
    # fila_inicio - 1 is reserved for column headers (written by _write_col_headers)
    for blank_row in range(5, fila_inicio - 1):
        ws.row_dimensions[blank_row].height = 15


# ---------------------------------------------------------------------------
# Column header row writer
# ---------------------------------------------------------------------------

def _write_col_headers(
    ws: Any,
    columnas: list[str],
    fila_inicio: int,
) -> None:
    """Write styled column headers in the row immediately before data rows.

    The column-header row is placed at ``fila_inicio - 1`` (1-based).

    Args:
        ws: An ``openpyxl`` ``Worksheet`` object.
        columnas: Ordered list of column header strings.
        fila_inicio: 1-based first data row index.
    """
    header_row = fila_inicio - 1  # 1-based
    ws.row_dimensions[header_row].height = 22

    for col_idx, col_name in enumerate(columnas, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        _apply_col_header_style(cell)

        # Auto-size column width based on header text length (min 10, max 30)
        col_letter = get_column_letter(col_idx)
        estimated_width = max(10, min(30, len(col_name) + 4))
        ws.column_dimensions[col_letter].width = estimated_width


# ---------------------------------------------------------------------------
# Instrucciones sheet writer
# ---------------------------------------------------------------------------

def _write_instrucciones_sheet(wb: Workbook, fila_inicio: int) -> None:
    """Add an 'Instrucciones' sheet with plain-text filling instructions.

    The sheet contains a title row and five numbered instructions that guide
    end users on how to fill in the template correctly.

    Args:
        wb: The ``openpyxl`` ``Workbook`` to add the sheet to.
        fila_inicio: 1-based data start row — embedded in instruction #2 text.
    """
    ws_instr = wb.create_sheet(title="Instrucciones")

    # Title
    title_cell = ws_instr["A1"]
    title_cell.value = "INSTRUCCIONES DE LLENADO"
    title_cell.font = Font(bold=True, size=13, color=_HEX_LABEL_TEXT, name="Calibri")
    title_cell.fill = PatternFill(fill_type="solid", fgColor=_HEX_LABEL_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_instr.row_dimensions[1].height = 24
    ws_instr.merge_cells("A1:E1")

    # Five numbered instructions
    instrucciones: list[str] = [
        "1. Complete los datos de contexto (UE, Meta, Año) en las celdas correspondientes.",
        f"2. Ingrese los datos a partir de la fila {fila_inicio}.",
        "3. No modifique las cabeceras de columna.",
        "4. Los campos de montos deben ser numéricos (sin formato de texto).",
        "5. No elimine ni agregue columnas.",
    ]

    for row_offset, text in enumerate(instrucciones, start=2):
        cell = ws_instr.cell(row=row_offset, column=1, value=text)
        cell.font = Font(size=10, name="Calibri", color="374151")
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        ws_instr.row_dimensions[row_offset].height = 18

    # Set a comfortable column width for readability
    ws_instr.column_dimensions["A"].width = 80


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_formato_catalog() -> list[dict[str, Any]]:
    """Return a safe copy of the complete format catalog.

    Each dict in the returned list contains the keys ``key``, ``nombre``,
    ``descripcion``, ``hoja``, ``columnas``, and ``fila_inicio``.

    Returns:
        Deep-copied list of all 10 format catalog entries.
    """
    return copy.deepcopy(FORMATO_CATALOG)


def generate_template(formato_key: str, output_path: Path) -> Path:
    """Generate a single Excel template file for the specified format.

    Creates a ``.xlsx`` workbook at ``output_path`` containing:

    * A data sheet named after the format's ``hoja`` value, with:
        - A four-row context header (title, UE, Meta, Año).
        - Blank filler rows (when ``fila_inicio > 5``).
        - Styled column headers at row ``fila_inicio - 1``.
    * An ``Instrucciones`` sheet with five numbered filling instructions.

    Args:
        formato_key: One of the 10 format keys defined in ``FORMATO_CATALOG``.
        output_path: Absolute ``Path`` where the ``.xlsx`` file will be saved.
                     Parent directories must already exist.

    Returns:
        The resolved ``output_path`` after the file has been written.

    Raises:
        KeyError: If ``formato_key`` is not found in the catalog.
        OSError: If ``output_path`` is not writable.
    """
    fmt = _CATALOG_BY_KEY.get(formato_key)
    if fmt is None:
        available = ", ".join(_CATALOG_BY_KEY.keys())
        raise KeyError(
            f"Formato key '{formato_key}' no encontrado en el catálogo. "
            f"Claves disponibles: {available}"
        )

    nombre: str = fmt["nombre"]
    hoja: str = fmt["hoja"]
    columnas: list[str] = fmt["columnas"]
    fila_inicio: int = fmt["fila_inicio"]
    num_cols: int = len(columnas)

    wb = Workbook()

    # Rename the default sheet to the format's sheet name
    ws = wb.active
    ws.title = hoja  # type: ignore[union-attr]

    # Freeze pane below the column-header row so users can scroll data easily
    freeze_row = fila_inicio  # first data row in 1-based; freeze pane is 0-based Excel address
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)  # type: ignore[union-attr]

    # Write context header block (rows 1-4 + optional filler)
    _write_context_header(ws, nombre, fila_inicio, num_cols)

    # Write styled column headers at row (fila_inicio - 1)
    _write_col_headers(ws, columnas, fila_inicio)

    # Add Instrucciones sheet
    _write_instrucciones_sheet(wb, fila_inicio)

    # Ensure parent directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(output_path))

    logger.info(
        "generate_template: key='%s' cols=%d fila_inicio=%d -> '%s'",
        formato_key,
        num_cols,
        fila_inicio,
        output_path,
    )

    return output_path.resolve()


def generate_all_templates(plantillas_dir: Path) -> list[str]:
    """Generate all 10 INEI format templates and save them to ``plantillas_dir``.

    File names follow the pattern ``plantilla_{formato_key}.xlsx`` so they can
    be served directly by the import router without any name-mapping step.

    Args:
        plantillas_dir: Directory where the ``.xlsx`` files will be written.
                        Created (including parents) if it does not exist.

    Returns:
        List of absolute file path strings for each generated template,
        in the same order as ``FORMATO_CATALOG``.

    Raises:
        OSError: If any individual file cannot be written to ``plantillas_dir``.
    """
    plantillas_dir.mkdir(parents=True, exist_ok=True)

    generated: list[str] = []
    for fmt in FORMATO_CATALOG:
        key: str = fmt["key"]
        file_path = plantillas_dir / f"plantilla_{key}.xlsx"
        resolved = generate_template(key, file_path)
        generated.append(str(resolved))

    logger.info(
        "generate_all_templates: %d templates written to '%s'",
        len(generated),
        plantillas_dir,
    )
    return generated
