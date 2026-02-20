"""Generate 10 complete example Excel files with realistic dummy data.

Each file mirrors the exact layout that the parsers expect:
- Context header rows (UE, Meta, Year) in the positions parsers scan
- Column headers at the expected row
- Data rows with valid values that will pass parser validation

Usage:
    cd backend
    py generate_example_data.py
"""
from __future__ import annotations

import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Resolve output directory
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "formatos" / "ejemplo"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Shared styling
BLUE_FILL = PatternFill(fill_type="solid", fgColor="3b82f6")
WHITE_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

# Dummy data pools
CLASIFICADORES = [
    "2.3.1.1.1.1", "2.3.1.1.1.2", "2.3.1.2.1.1", "2.3.1.2.1.2",
    "2.3.1.3.1.1", "2.3.1.5.1.1", "2.3.1.5.1.2", "2.3.1.5.3.1",
    "2.3.1.9.1.1", "2.3.1.9.1.2", "2.3.2.1.2.1", "2.3.2.1.2.2",
    "2.3.2.2.1.1", "2.3.2.2.2.1", "2.3.2.2.4.4", "2.6.3.2.3.1",
]
DESCRIPCIONES_GASTO = [
    "Alimentos y bebidas para consumo humano",
    "Vestuario, accesorios y prendas diversas",
    "Combustibles, carburantes, lubricantes",
    "Papeleria en general, utiles de oficina",
    "Aseo, limpieza y tocador",
    "Materiales de construccion",
    "Repuestos y accesorios",
    "Suministros para mantenimiento",
    "Material de escritorio",
    "Libros, textos y otros materiales impresos",
    "Servicio de suministro de energia electrica",
    "Servicio de agua y desague",
    "Servicio de telefonia movil",
    "Servicio de internet",
    "Servicio de mensajeria",
    "Equipos computacionales y perifericos",
]
AO_CODES = [
    "AOI.01.01", "AOI.01.02", "AOI.01.03", "AOI.02.01", "AOI.02.02",
    "AOI.03.01", "AOI.03.02", "AOI.04.01", "AOI.04.02", "AOI.05.01",
]
AO_NAMES = [
    "Gestion y seguimiento de actividades estadisticas",
    "Supervision de operaciones de campo",
    "Procesamiento de datos estadisticos",
    "Produccion de indicadores economicos",
    "Elaboracion de informes tecnicos",
    "Capacitacion al personal de campo",
    "Mantenimiento de infraestructura TI",
    "Administracion de recursos humanos",
    "Gestion logistica y adquisiciones",
    "Difusion de resultados estadisticos",
]
METAS = ["0001", "0002", "0003", "0004", "0005"]
META_DESCS = [
    "Produccion estadistica nacional",
    "Censos y encuestas nacionales",
    "Infraestructura estadistica",
    "Difusion y comunicacion institucional",
    "Gestion administrativa",
]
TAREAS = ["T001", "T002", "T003", "T004", "T005"]
TAREA_DESCS = [
    "Recopilacion de informacion primaria",
    "Validacion y consistencia de datos",
    "Procesamiento y tabulacion",
    "Elaboracion de cuadros estadisticos",
    "Revision y control de calidad",
]
NOMBRES_RRHH = [
    ("Garcia Lopez, Juan Carlos", "12345678"),
    ("Rodriguez Perez, Maria Elena", "23456789"),
    ("Fernandez Torres, Pedro Luis", "34567890"),
    ("Martinez Diaz, Ana Patricia", "45678901"),
    ("Sanchez Ramos, Jose Miguel", "56789012"),
    ("Vargas Quispe, Rosa Maria", "67890123"),
    ("Mendoza Flores, Carlos Alberto", "78901234"),
    ("Castillo Rivera, Luz Marina", "89012345"),
    ("Rojas Huaman, Luis Fernando", "90123456"),
    ("Torres Gomez, Sandra Paola", "01234567"),
]
CARGOS = [
    "Especialista Estadistico", "Analista de Datos", "Tecnico de Campo",
    "Jefe de Area", "Asistente Administrativo", "Coordinador Regional",
]
AREAS = ["OTIN", "DEC", "OTA", "OTPP", "OTD", "OGPP"]
REGIMENES = ["276", "728", "1057 (CAS)", "Locacion de Servicios"]
TIPOS_CONTRATO = ["Indeterminado", "Plazo Fijo", "CAS", "Orden de Servicio"]


def _styled_header(ws, row, cols):
    for col_idx, name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _context_rows(ws, ue="001 - INEI SEDE CENTRAL", meta="0001", anio=2026):
    """Write standard 4-row context block."""
    ws["A1"] = "INSTITUTO NACIONAL DE ESTADISTICA E INFORMATICA"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Unidad Ejecutora:"
    ws["B2"] = ue
    ws["A3"] = "Meta Presupuestal:"
    ws["B3"] = meta
    ws["A4"] = "Ano Fiscal:"
    ws["B4"] = anio


def _random_monthly(total):
    """Split total into 12 random monthly values that sum to total."""
    if total <= 0:
        return [0.0] * 12
    weights = [random.uniform(0.5, 1.5) for _ in range(12)]
    s = sum(weights)
    monthly = [round(total * w / s, 2) for w in weights]
    # Adjust last month to ensure exact sum
    diff = round(total - sum(monthly), 2)
    monthly[11] = round(monthly[11] + diff, 2)
    return monthly


def gen_cuadro_ao_meta():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuadro AO-Meta"
    # Cuadro AO-Meta has a simpler layout: direct header at row 1
    # matching what real INEI files look like for master data
    cols = [
        "Codigo UE", "Nombre UE", "Sigla",
        "Codigo Meta", "Sec. Funcional", "Descripcion Meta",
        "Codigo AO", "Nombre AO", "OEI", "AEI",
    ]
    _styled_header(ws, 1, cols)
    for i in range(10):
        r = 2 + i
        ws.cell(row=r, column=1, value="001")
        ws.cell(row=r, column=2, value="INEI SEDE CENTRAL")
        ws.cell(row=r, column=3, value="INEI")
        ws.cell(row=r, column=4, value=METAS[i % 5])
        ws.cell(row=r, column=5, value=f"00{i+1}")
        ws.cell(row=r, column=6, value=META_DESCS[i % 5])
        ws.cell(row=r, column=7, value=AO_CODES[i])
        ws.cell(row=r, column=8, value=AO_NAMES[i])
        ws.cell(row=r, column=9, value=f"OEI.0{(i%3)+1}")
        ws.cell(row=r, column=10, value=f"AEI.0{(i%4)+1}")
    wb.save(str(OUTPUT_DIR / "ejemplo_cuadro_ao_meta.xlsx"))
    print("  cuadro_ao_meta OK")


def gen_tablas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tablas"
    # Tablas parser expects header at row 1 with Clasificador + Descripcion + Tipo Generico
    cols = ["Clasificador", "Descripcion", "Tipo Generico"]
    _styled_header(ws, 1, cols)
    genericos = ["2.3", "2.6", "2.1"]
    for i, clas in enumerate(CLASIFICADORES):
        r = 2 + i
        ws.cell(row=r, column=1, value=clas)
        ws.cell(row=r, column=2, value=DESCRIPCIONES_GASTO[i])
        ws.cell(row=r, column=3, value=random.choice(genericos))
    wb.save(str(OUTPUT_DIR / "ejemplo_tablas.xlsx"))
    print("  tablas OK")


def gen_formato1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 1"
    _context_rows(ws)
    ws["A5"] = ""
    ws["A6"] = ""
    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    cols = ["Clasificador", "Descripcion", "PIA", "PIM"] + months + ["Total"]
    _styled_header(ws, 7, cols)
    for i in range(12):
        r = 8 + i
        pim = round(random.uniform(50000, 500000), 2)
        pia = round(pim * random.uniform(0.8, 1.0), 2)
        monthly = _random_monthly(pim)
        ws.cell(row=r, column=1, value=CLASIFICADORES[i])
        ws.cell(row=r, column=2, value=DESCRIPCIONES_GASTO[i])
        ws.cell(row=r, column=3, value=pia)
        ws.cell(row=r, column=4, value=pim)
        for m_idx, mval in enumerate(monthly):
            ws.cell(row=r, column=5 + m_idx, value=mval)
        ws.cell(row=r, column=17, value=round(sum(monthly), 2))
    wb.save(str(OUTPUT_DIR / "ejemplo_formato1.xlsx"))
    print("  formato1 OK")


def gen_formato2():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 2"
    _context_rows(ws)
    ws["A5"] = ""
    ws["A6"] = ""
    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    cols = [
        "Cod Meta", "Desc Meta", "Cod AO", "Desc AO",
        "Cod Tarea", "Desc Tarea", "Clasificador", "Desc Clasificador", "PIM",
    ] + months
    _styled_header(ws, 7, cols)
    row = 8
    for t in range(10):
        meta_idx = t % 5
        ao_idx = t % 10
        tarea_idx = t % 5
        pim = round(random.uniform(20000, 200000), 2)
        monthly = _random_monthly(pim)
        ws.cell(row=row, column=1, value=METAS[meta_idx])
        ws.cell(row=row, column=2, value=META_DESCS[meta_idx])
        ws.cell(row=row, column=3, value=AO_CODES[ao_idx])
        ws.cell(row=row, column=4, value=AO_NAMES[ao_idx])
        ws.cell(row=row, column=5, value=TAREAS[tarea_idx])
        ws.cell(row=row, column=6, value=TAREA_DESCS[tarea_idx])
        ws.cell(row=row, column=7, value=CLASIFICADORES[t % 16])
        ws.cell(row=row, column=8, value=DESCRIPCIONES_GASTO[t % 16])
        ws.cell(row=row, column=9, value=pim)
        for m_idx, mval in enumerate(monthly):
            ws.cell(row=row, column=10 + m_idx, value=mval)
        row += 1
    wb.save(str(OUTPUT_DIR / "ejemplo_formato2.xlsx"))
    print("  formato2 OK")


def gen_formato3():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 3"
    _context_rows(ws)
    ws["A5"] = ""
    ws["A6"] = ""
    cols = [
        "Cod Meta", "Desc Meta", "Cod AO", "Desc AO",
        "Cod Tarea", "Desc Tarea", "Clasificador", "Desc Clasificador",
        "PIM", "Programado", "Ejecutado", "Saldo", "% Avance",
        "Justificacion", "Observaciones",
    ]
    _styled_header(ws, 7, cols)
    justificaciones = [
        "Ejecucion conforme al cronograma establecido",
        "Retraso por demora en proceso de seleccion",
        "Adelanto de actividades por necesidad institucional",
        "Pendiente conformidad de area usuaria",
        "En proceso de adquisicion de bienes",
    ]
    row = 8
    for t in range(10):
        pim = round(random.uniform(30000, 300000), 2)
        ejecutado = round(pim * random.uniform(0.3, 0.95), 2)
        programado = round(pim * random.uniform(0.8, 1.0), 2)
        saldo = round(pim - ejecutado, 2)
        avance = round(ejecutado / pim * 100, 2) if pim > 0 else 0
        ws.cell(row=row, column=1, value=METAS[t % 5])
        ws.cell(row=row, column=2, value=META_DESCS[t % 5])
        ws.cell(row=row, column=3, value=AO_CODES[t % 10])
        ws.cell(row=row, column=4, value=AO_NAMES[t % 10])
        ws.cell(row=row, column=5, value=TAREAS[t % 5])
        ws.cell(row=row, column=6, value=TAREA_DESCS[t % 5])
        ws.cell(row=row, column=7, value=CLASIFICADORES[t % 16])
        ws.cell(row=row, column=8, value=DESCRIPCIONES_GASTO[t % 16])
        ws.cell(row=row, column=9, value=pim)
        ws.cell(row=row, column=10, value=programado)
        ws.cell(row=row, column=11, value=ejecutado)
        ws.cell(row=row, column=12, value=saldo)
        ws.cell(row=row, column=13, value=avance)
        ws.cell(row=row, column=14, value=random.choice(justificaciones))
        ws.cell(row=row, column=15, value="Sin observaciones" if avance > 70 else "Requiere atencion")
        row += 1
    wb.save(str(OUTPUT_DIR / "ejemplo_formato3.xlsx"))
    print("  formato3 OK")


def gen_formato04():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 04"
    ws["A1"] = "NOTA DE MODIFICACION PRESUPUESTAL"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Unidad Ejecutora:"
    ws["C2"] = "001 - INEI SEDE CENTRAL"
    ws["A3"] = "Nota de Modificacion:"
    ws["C3"] = "NM-2026-001"
    ws["A4"] = "Fecha:"
    ws["F4"] = "15/02/2026"
    ws["F3"] = ""
    ws["F4"] = 2026
    ws["A5"] = ""
    ws["A6"] = ""
    cols = ["Clasificador", "Descripcion", "Asignado", "Habilitadora", "Habilitada", "PIM Resultante"]
    _styled_header(ws, 7, cols)
    for i in range(8):
        r = 8 + i
        asignado = round(random.uniform(50000, 300000), 2)
        # Half rows are habilitadoras, half habilitadas
        if i % 2 == 0:
            hab_r = round(random.uniform(10000, 50000), 2)
            hab_g = 0.0
        else:
            hab_r = 0.0
            hab_g = round(random.uniform(10000, 50000), 2)
        pim_res = round(asignado + hab_r - hab_g, 2)
        ws.cell(row=r, column=1, value=CLASIFICADORES[i])
        ws.cell(row=r, column=2, value=DESCRIPCIONES_GASTO[i])
        ws.cell(row=r, column=3, value=asignado)
        ws.cell(row=r, column=4, value=hab_r)
        ws.cell(row=r, column=5, value=hab_g)
        ws.cell(row=r, column=6, value=pim_res)
    wb.save(str(OUTPUT_DIR / "ejemplo_formato04.xlsx"))
    print("  formato04 OK")


def gen_formato5a():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 5.A"
    # Context rows (Formato 5.A has context up to row ~10)
    ws["A1"] = "FORMATO 5.A - PROGRAMACION MENSUAL DE ACTIVIDADES OPERATIVAS"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ""
    ws["A3"] = "Unidad Ejecutora:"
    ws["C3"] = "001 - INEI SEDE CENTRAL"
    ws["F3"] = ""
    ws["A4"] = "Meta Presupuestal:"
    ws["F4"] = "0001"
    ws["A5"] = "Ano Fiscal:"
    ws["F5"] = 2026
    for r in range(6, 10):
        ws.cell(row=r, column=1, value="")
    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    cols = ["Codigo AO", "Nombre AO"] + months + ["Total Programado"]
    _styled_header(ws, 11, cols)
    for i in range(10):
        r = 12 + i
        total = round(random.uniform(80000, 600000), 2)
        monthly = _random_monthly(total)
        ws.cell(row=r, column=1, value=AO_CODES[i])
        ws.cell(row=r, column=2, value=AO_NAMES[i])
        for m_idx, mval in enumerate(monthly):
            ws.cell(row=r, column=3 + m_idx, value=mval)
        ws.cell(row=r, column=15, value=round(sum(monthly), 2))
    wb.save(str(OUTPUT_DIR / "ejemplo_formato5a.xlsx"))
    print("  formato5a OK")


def gen_formato5b():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 5.B"
    ws["A1"] = "FORMATO 5.B - EJECUCION MENSUAL DE ACTIVIDADES OPERATIVAS"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Unidad Ejecutora:"
    ws["C3"] = "001 - INEI SEDE CENTRAL"
    ws["A4"] = ""
    ws["F4"] = "0001"
    ws["A5"] = ""
    ws["F5"] = 2026
    months_full = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    # Row 9 (1-based): Two-row compound header
    # Row 9: "Codigo AO", "Nombre AO", then month names (each spanning 3 cols)
    ws.cell(row=9, column=1, value="Codigo AO")
    ws.cell(row=9, column=2, value="Nombre AO")
    col = 3
    for month_name in months_full:
        ws.cell(row=9, column=col, value=month_name)
        # Merge across 3 columns for each month
        ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col + 2)
        col += 3
    # Add Total columns
    ws.cell(row=9, column=col, value="Total")
    ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col + 2)

    # Row 10: sub-headers (Programado, Ejecutado, Saldo) repeated for each month
    ws.cell(row=10, column=1, value="")
    ws.cell(row=10, column=2, value="")
    col = 3
    for _ in range(13):  # 12 months + 1 total
        ws.cell(row=10, column=col, value="Programado")
        ws.cell(row=10, column=col + 1, value="Ejecutado")
        ws.cell(row=10, column=col + 2, value="Saldo")
        col += 3

    # Style header rows
    for r in (9, 10):
        for c in range(1, col):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                cell.font = WHITE_FONT
                cell.fill = BLUE_FILL
                cell.alignment = HEADER_ALIGN

    # Data from row 12
    for i in range(10):
        r = 12 + i
        pim = round(random.uniform(100000, 800000), 2)
        prog_monthly = _random_monthly(pim)
        ejec_monthly = [round(p * random.uniform(0.5, 1.0), 2) for p in prog_monthly]
        saldo_monthly = [round(p - e, 2) for p, e in zip(prog_monthly, ejec_monthly)]
        total_prog = round(sum(prog_monthly), 2)
        total_ejec = round(sum(ejec_monthly), 2)
        total_saldo = round(total_prog - total_ejec, 2)

        ws.cell(row=r, column=1, value=AO_CODES[i])
        ws.cell(row=r, column=2, value=AO_NAMES[i])
        col = 3
        for m in range(12):
            ws.cell(row=r, column=col, value=prog_monthly[m])
            ws.cell(row=r, column=col + 1, value=ejec_monthly[m])
            ws.cell(row=r, column=col + 2, value=saldo_monthly[m])
            col += 3
        # Totals
        ws.cell(row=r, column=col, value=total_prog)
        ws.cell(row=r, column=col + 1, value=total_ejec)
        ws.cell(row=r, column=col + 2, value=total_saldo)
    wb.save(str(OUTPUT_DIR / "ejemplo_formato5b.xlsx"))
    print("  formato5b OK")


def gen_formato5_resumen():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 5 Resumen"
    ws["A1"] = "FORMATO 5 RESUMEN - EJECUCION POR ACTIVIDAD OPERATIVA"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Unidad Ejecutora:"
    ws["C2"] = "001 - INEI SEDE CENTRAL"
    ws["A3"] = "Meta Presupuestal:"
    ws["F3"] = "0001"
    ws["A4"] = "Ano Fiscal:"
    ws["F4"] = 2026
    ws["A5"] = ""
    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    cols = [
        "Codigo AO", "Nombre AO", "PIM", "CCP", "Compromiso Anual",
        "Devengado", "Girado", "Saldo", "% Avance PIM", "% Avance CCP", "Semaforo",
    ] + months
    _styled_header(ws, 6, cols)
    semaforos = ["VERDE", "AMARILLO", "ROJO"]
    for i in range(10):
        r = 7 + i
        pim = round(random.uniform(100000, 700000), 2)
        ccp = round(pim * random.uniform(0.7, 1.0), 2)
        compromiso = round(ccp * random.uniform(0.8, 1.0), 2)
        devengado = round(compromiso * random.uniform(0.6, 0.95), 2)
        girado = round(devengado * random.uniform(0.9, 1.0), 2)
        saldo = round(pim - devengado, 2)
        pct_pim = round(devengado / pim * 100, 2) if pim > 0 else 0
        pct_ccp = round(devengado / ccp * 100, 2) if ccp > 0 else 0
        semaforo = "VERDE" if pct_pim >= 90 else ("AMARILLO" if pct_pim >= 70 else "ROJO")
        dev_monthly = _random_monthly(devengado)

        ws.cell(row=r, column=1, value=AO_CODES[i])
        ws.cell(row=r, column=2, value=AO_NAMES[i])
        ws.cell(row=r, column=3, value=pim)
        ws.cell(row=r, column=4, value=ccp)
        ws.cell(row=r, column=5, value=compromiso)
        ws.cell(row=r, column=6, value=devengado)
        ws.cell(row=r, column=7, value=girado)
        ws.cell(row=r, column=8, value=saldo)
        ws.cell(row=r, column=9, value=pct_pim)
        ws.cell(row=r, column=10, value=pct_ccp)
        ws.cell(row=r, column=11, value=semaforo)
        for m_idx, mval in enumerate(dev_monthly):
            ws.cell(row=r, column=12 + m_idx, value=mval)
    wb.save(str(OUTPUT_DIR / "ejemplo_formato5_resumen.xlsx"))
    print("  formato5_resumen OK")


def gen_anexo01():
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo 01"
    _context_rows(ws)
    ws["A5"] = ""
    ws["A6"] = ""
    cols = [
        "N", "DNI", "Apellidos y Nombres", "Cargo", "Area",
        "Regimen Laboral", "Tipo Contrato", "Fecha Inicio", "Fecha Fin",
        "Remuneracion Mensual", "Observaciones", "Estado",
    ]
    _styled_header(ws, 7, cols)
    for i, (nombre, dni) in enumerate(NOMBRES_RRHH):
        r = 8 + i
        remu = round(random.uniform(2500, 12000), 2)
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=dni)
        ws.cell(row=r, column=3, value=nombre)
        ws.cell(row=r, column=4, value=random.choice(CARGOS))
        ws.cell(row=r, column=5, value=random.choice(AREAS))
        ws.cell(row=r, column=6, value=random.choice(REGIMENES))
        ws.cell(row=r, column=7, value=random.choice(TIPOS_CONTRATO))
        ws.cell(row=r, column=8, value="01/01/2026")
        ws.cell(row=r, column=9, value="31/12/2026")
        ws.cell(row=r, column=10, value=remu)
        ws.cell(row=r, column=11, value="")
        ws.cell(row=r, column=12, value="ACTIVO")
    wb.save(str(OUTPUT_DIR / "ejemplo_anexo01.xlsx"))
    print("  anexo01 OK")


def main():
    random.seed(42)  # Reproducible
    print(f"Generating example files in: {OUTPUT_DIR}")
    gen_cuadro_ao_meta()
    gen_tablas()
    gen_formato1()
    gen_formato2()
    gen_formato3()
    gen_formato04()
    gen_formato5a()
    gen_formato5b()
    gen_formato5_resumen()
    gen_anexo01()
    print(f"\nDone! {len(list(OUTPUT_DIR.glob('*.xlsx')))} files generated.")


if __name__ == "__main__":
    main()
