"""Generate example Excel files for ALL 12 INEI formats with consistent data.

All example files share the same UE codes, Meta codes, Clasificador codes, and
AO CEPLAN codes so that importing them in order produces a complete, working
dashboard with realistic numbers.

Also generates plantillas (templates) for SIAF and SIGA, and seed data for
contratos menores and adquisiciones.

Usage:
    py generate_examples.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).parent
EJEMPLOS_DIR = BASE_DIR / "formatos" / "ejemplo"
PLANTILLAS_DIR = BASE_DIR / "formatos" / "plantillas"

# =========================================================================
# Styles
# =========================================================================
_BLUE = "3b82f6"
_DARK = "1E3A5F"
_LIGHT_BG = "EFF6FF"
_BORDER_COLOR = "CBD5E1"

thin_side = Side(style="thin", color=_BORDER_COLOR)
THIN_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
HEADER_FILL = PatternFill(fill_type="solid", fgColor=_BLUE)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
TITLE_FILL = PatternFill(fill_type="solid", fgColor=_DARK)

LABEL_FONT = Font(bold=True, color=_DARK, size=10, name="Calibri")
LABEL_FILL = PatternFill(fill_type="solid", fgColor=_LIGHT_BG)

DATA_FONT = Font(size=10, name="Calibri")


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def style_title(cell):
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_label(cell):
    cell.font = LABEL_FONT
    cell.fill = LABEL_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right", vertical="center")


def style_data(cell):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER


def write_context(ws, title, contexts, num_cols, *, safe_labels=False):
    """Write title and context rows. If safe_labels=True, use abbreviated
    labels that don't conflict with parser keyword detection."""
    from openpyxl.utils import get_column_letter
    ws["A1"] = title
    style_title(ws["A1"])
    end_col = get_column_letter(min(num_cols, 6))
    if num_cols > 1:
        ws.merge_cells(f"A1:{end_col}1")
    ws.row_dimensions[1].height = 28

    # Safe label mapping: avoid words like "meta", "clasificador", "tarea",
    # "justificacion", "codigo", "nombre", "ceplan", "devengado", "girado", etc.
    _SAFE = {
        "Unidad Ejecutora:": "UE:",
        "Meta Presupuestal:": "MP:",
        "Ano Fiscal:": "Ejercicio:",
        "Entidad:": "Entidad:",
        "Fecha:": "Fecha:",
        "Ejercicio:": "Ejercicio:",
        "Periodo:": "Periodo:",
        "Nota Modificacion:": "Nota Mod.:",
    }
    for i, (label, value) in enumerate(contexts, start=2):
        display_label = _SAFE.get(label, label) if safe_labels else label
        lbl_cell = ws.cell(row=i, column=1, value=display_label)
        style_label(lbl_cell)
        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = DATA_FONT
        val_cell.border = THIN_BORDER
        ws.row_dimensions[i].height = 18


def write_headers_and_data(ws, headers, data, header_row):
    from openpyxl.utils import get_column_letter
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(h)) + 4)
    ws.row_dimensions[header_row].height = 22
    for row_offset, row_data in enumerate(data, start=header_row + 1):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            style_data(cell)


def fmt_money(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in cols:
            ws.cell(row=r, column=c).number_format = '#,##0.00'


# =========================================================================
# Shared Constants — consistent across ALL formats
# =========================================================================
UE_CODIGO = "001"
UE_NOMBRE = "001 - INEI SEDE CENTRAL"
META_CODIGO = "0001"
META_DESC = "Gestion Administrativa"
ANIO = 2026

# 10 AO codes used across cuadro_ao_meta, formato5a, formato5b, formato5_resumen
AO_DATA = [
    ("AO0101001", "Gestion de Tecnologias de Informacion", "OEI.01", "AEI.01.01"),
    ("AO0101002", "Desarrollo de Sistemas de Informacion", "OEI.01", "AEI.01.01"),
    ("AO0102001", "Gestion de Infraestructura Tecnologica", "OEI.01", "AEI.01.02"),
    ("AO0201001", "Censos y Encuestas Nacionales", "OEI.02", "AEI.02.01"),
    ("AO0201002", "Indicadores Economicos y Sociales", "OEI.02", "AEI.02.01"),
    ("AO0202001", "Publicaciones Estadisticas", "OEI.02", "AEI.02.02"),
    ("AO0301001", "Gestion de Recursos Humanos", "OEI.03", "AEI.03.01"),
    ("AO0301002", "Capacitacion y Desarrollo", "OEI.03", "AEI.03.01"),
    ("AO0302001", "Gestion Presupuestal y Financiera", "OEI.03", "AEI.03.02"),
    ("AO0302002", "Gestion Logistica y Patrimonial", "OEI.03", "AEI.03.02"),
]

# 12 clasificadores used across tablas, formato1, formato04, siaf
CLASIFICADORES = [
    ("2.1.1.1.1.1", "2.1", "FUNCIONARIOS ELEGIDOS"),
    ("2.1.1.1.1.2", "2.1", "PERSONAL ADMINISTRATIVO NOMBRADO"),
    ("2.3.1.1.1.1", "2.3", "ALIMENTOS Y BEBIDAS PARA CONSUMO HUMANO"),
    ("2.3.1.5.1.1", "2.3", "MATERIALES Y UTILES DE OFICINA"),
    ("2.3.1.5.1.2", "2.3", "PAPELERIA EN GENERAL"),
    ("2.3.2.1.2.1", "2.3", "PASAJES Y GASTOS DE TRANSPORTE"),
    ("2.3.2.2.1.1", "2.3", "SERVICIO DE ENERGIA ELECTRICA"),
    ("2.3.2.2.2.1", "2.3", "SERVICIO DE AGUA Y DESAGUE"),
    ("2.3.2.2.3.1", "2.3", "SERVICIO DE TELEFONIA FIJA"),
    ("2.3.2.7.1.1", "2.3", "SERVICIOS DE CONSULTORIA"),
    ("2.3.2.7.11.99", "2.3", "OTROS SERVICIOS DIVERSOS"),
    ("2.6.3.2.1.1", "2.6", "EQUIPOS COMPUTACIONALES Y PERIFERICOS"),
]

# Monthly budget amounts (Jan-Dec) for Formato 1
MONTHLY_BUDGETS = {
    "2.1.1.1.1.1":   [45000, 45000, 45000, 45000, 45000, 45000, 45000, 45000, 45000, 45000, 45000, 45000],
    "2.1.1.1.1.2":   [32000, 32000, 32000, 32000, 32000, 32000, 32000, 32000, 32000, 32000, 32000, 32000],
    "2.3.1.1.1.1":   [2500, 2500, 2500, 2500, 2500, 2500, 2500, 2500, 2500, 2500, 2500, 2500],
    "2.3.1.5.1.1":   [15000, 12000, 14000, 13000, 15000, 12000, 14000, 13000, 15000, 12000, 14000, 13000],
    "2.3.1.5.1.2":   [7500, 7000, 7500, 7000, 7500, 7000, 7500, 7000, 7500, 7000, 7500, 7000],
    "2.3.2.1.2.1":   [12000, 10000, 8000, 12000, 15000, 10000, 8000, 12000, 15000, 10000, 8000, 12000],
    "2.3.2.2.1.1":   [20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000],
    "2.3.2.2.2.1":   [4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000, 4000],
    "2.3.2.2.3.1":   [3200, 3200, 3200, 3200, 3200, 3200, 3200, 3200, 3200, 3200, 3200, 3200],
    "2.3.2.7.1.1":   [50000, 40000, 45000, 50000, 40000, 45000, 50000, 40000, 45000, 50000, 40000, 45000],
    "2.3.2.7.11.99": [8000, 6000, 7000, 8000, 6000, 7000, 8000, 6000, 7000, 8000, 6000, 7000],
    "2.6.3.2.1.1":   [0, 0, 50000, 0, 0, 50000, 0, 0, 50000, 0, 0, 40000],
}

# AO monthly programado (Jan-Dec) for Formato 5A/5B
AO_MONTHLY_PROG = {
    "AO0101001": [25000, 22000, 28000, 25000, 30000, 25000, 28000, 25000, 30000, 25000, 28000, 25000],
    "AO0101002": [18000, 15000, 20000, 18000, 22000, 18000, 20000, 18000, 22000, 18000, 20000, 18000],
    "AO0102001": [12000, 12000, 15000, 12000, 12000, 15000, 12000, 12000, 15000, 12000, 12000, 15000],
    "AO0201001": [40000, 35000, 42000, 40000, 45000, 40000, 42000, 40000, 45000, 40000, 42000, 40000],
    "AO0201002": [15000, 12000, 15000, 15000, 18000, 15000, 15000, 15000, 18000, 15000, 15000, 15000],
    "AO0202001": [8000, 8000, 10000, 8000, 8000, 10000, 8000, 8000, 10000, 8000, 8000, 10000],
    "AO0301001": [20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000],
    "AO0301002": [10000, 8000, 12000, 10000, 8000, 12000, 10000, 8000, 12000, 10000, 8000, 12000],
    "AO0302001": [35000, 30000, 35000, 35000, 38000, 35000, 35000, 35000, 38000, 35000, 35000, 35000],
    "AO0302002": [15000, 12000, 15000, 15000, 18000, 15000, 15000, 15000, 18000, 15000, 15000, 15000],
}

# Execution ratios by month (Feb is current month 2026, so execute ~80% for past months)
EXEC_RATIOS = [0.95, 0.85, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # Only Jan-Feb have execution


# =========================================================================
# 1. CUADRO AO-META (Master data)
# =========================================================================
def generate_cuadro_ao_meta():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuadro AO-Meta"

    # Title and context must NOT contain parser keywords: "codigo", "nombre", "ceplan", "meta"
    write_context(ws, "Datos Maestros AO", [
        ("UE:", UE_NOMBRE),
        ("MP:", META_CODIGO),
        ("Ejercicio:", ANIO),
    ], 10)

    # Header row at row 6 (data starts at row 7 per catalog)
    headers = [
        "Codigo UE", "Nombre UE", "Sigla", "Tipo",
        "Codigo Meta", "Sec. Funcional", "Descripcion Meta",
        "Codigo AO", "Nombre AO", "OEI", "AEI",
    ]
    data = []
    for ao_code, ao_name, oei, aei in AO_DATA:
        data.append((
            UE_CODIGO, "INEI SEDE CENTRAL", "INEI-CENTRAL", "CENTRAL",
            META_CODIGO, "0001", META_DESC,
            ao_code, ao_name, oei, aei,
        ))

    write_headers_and_data(ws, headers, data, header_row=6)
    path = EJEMPLOS_DIR / "ejemplo_cuadro_ao_meta.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} AOs)")


# =========================================================================
# 2. TABLAS (Clasificadores de Gasto)
# =========================================================================
def generate_tablas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tablas"

    # Title must NOT contain "clasificador", "tipo generico" etc.
    write_context(ws, "Tablas de Referencia", [
        ("Entidad:", "INEI"),
        ("Fecha:", "2026"),
    ], 6)

    headers = ["Clasificador", "Tipo Generico", "Tipo Especifico", "Sub Tipo", "Descripcion", "Estado"]
    data = []
    for cod, tipo_gen, desc in CLASIFICADORES:
        parts = cod.split(".")
        tipo_esp = f"{parts[0]}.{parts[1]}.{parts[2]}" if len(parts) >= 3 else tipo_gen
        sub_tipo = f"{parts[0]}.{parts[1]}.{parts[2]}.{parts[3]}" if len(parts) >= 4 else tipo_esp
        data.append((cod, tipo_gen, tipo_esp, sub_tipo, desc, "ACTIVO"))

    write_headers_and_data(ws, headers, data, header_row=5)
    path = EJEMPLOS_DIR / "ejemplo_tablas.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} clasificadores)")


# =========================================================================
# 3. FORMATO 1 — Programacion Presupuestal Anual
# =========================================================================
def generate_formato1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 1"

    # Context rows matching parser expectations:
    # Row 1: Title
    # Row 2 (0-based 1): col B = UE nombre
    # Row 3 (0-based 2): col B = Meta codigo
    # Row 4 (0-based 3): col B = Ano
    write_context(ws, "Formato 1 - Programacion Presupuestal Anual", [
        ("Unidad Ejecutora:", UE_NOMBRE),
        ("Meta Presupuestal:", META_CODIGO),
        ("Ano Fiscal:", ANIO),
    ], 17)

    # Row 5-6: blank, Row 7: headers, Row 8+: data
    headers = [
        "Clasificador", "Descripcion", "PIA", "PIM",
        "Ene", "Feb", "Mar", "Abr", "May", "Jun",
        "Jul", "Ago", "Sep", "Oct", "Nov", "Dic", "Total",
    ]

    data = []
    for cod, _, desc in CLASIFICADORES:
        monthly = MONTHLY_BUDGETS.get(cod, [0]*12)
        total = sum(monthly)
        pia = int(total * 0.95)  # PIA slightly less than PIM
        pim = total
        row = [cod, desc, pia, pim] + monthly + [total]
        data.append(tuple(row))

    write_headers_and_data(ws, headers, data, header_row=7)
    fmt_money(ws, 8, 7 + len(data), list(range(3, 18)))  # cols C-Q

    path = EJEMPLOS_DIR / "ejemplo_formato1.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} clasificadores)")


# =========================================================================
# 4. FORMATO 2 — Programacion por Tareas
# =========================================================================
def generate_formato2():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 2"

    # Title must NOT contain "tarea" or "clasificador" — parser keywords
    # Context positions: (2,1)=ue_nombre, (2,3)=ue_codigo, (3,3)=meta_codigo, (4,3)=anio
    ws["A1"] = "FORMATO 2"
    style_title(ws["A1"])
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28
    # Row 2: blank
    # Row 3: UE context — parser reads (2,1) and (2,3)
    ws.cell(row=3, column=1, value="UE:").font = LABEL_FONT
    ws.cell(row=3, column=2, value=UE_NOMBRE).font = DATA_FONT
    ws.cell(row=3, column=4, value=UE_CODIGO).font = DATA_FONT
    # Row 4: Meta — parser reads (3,3)
    ws.cell(row=4, column=1, value="MP:").font = LABEL_FONT
    ws.cell(row=4, column=4, value=META_CODIGO).font = DATA_FONT
    # Row 5: Anio — parser reads (4,3)
    ws.cell(row=5, column=1, value="Ejercicio:").font = LABEL_FONT
    ws.cell(row=5, column=4, value=ANIO).font = DATA_FONT

    headers = [
        "Cod Meta", "Desc Meta", "Cod AO", "Desc AO",
        "Cod Tarea", "Desc Tarea", "Clasificador", "Desc Clasificador", "PIM",
        "Ene", "Feb", "Mar", "Abr", "May", "Jun",
        "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
    ]

    data = []
    task_n = 1
    for ao_code, ao_name, _, _ in AO_DATA[:5]:
        for cod, _, desc in CLASIFICADORES[:3]:
            monthly = [m // 5 for m in MONTHLY_BUDGETS.get(cod, [0]*12)]
            pim = sum(monthly)
            row = (
                META_CODIGO, META_DESC, ao_code, ao_name,
                f"T{task_n:03d}", f"Actividad {task_n}", cod, desc, pim,
                *monthly,
            )
            data.append(row)
            task_n += 1

    write_headers_and_data(ws, headers, data, header_row=7)
    path = EJEMPLOS_DIR / "ejemplo_formato2.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} tareas)")


# =========================================================================
# 5. FORMATO 3 — Tareas con Justificacion
# =========================================================================
def generate_formato3():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 3"

    # Title must NOT contain "tarea" or "justificacion" — parser keywords
    # Context positions same as Formato 2
    ws["A1"] = "FORMATO 3"
    style_title(ws["A1"])
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28
    ws.cell(row=3, column=1, value="UE:").font = LABEL_FONT
    ws.cell(row=3, column=2, value=UE_NOMBRE).font = DATA_FONT
    ws.cell(row=3, column=4, value=UE_CODIGO).font = DATA_FONT
    ws.cell(row=4, column=1, value="MP:").font = LABEL_FONT
    ws.cell(row=4, column=4, value=META_CODIGO).font = DATA_FONT
    ws.cell(row=5, column=1, value="Ejercicio:").font = LABEL_FONT
    ws.cell(row=5, column=4, value=ANIO).font = DATA_FONT

    headers = [
        "Cod Meta", "Desc Meta", "Cod AO", "Desc AO",
        "Cod Tarea", "Desc Tarea", "Clasificador", "Desc Clasificador",
        "PIM", "Programado", "Ejecutado", "Saldo", "% Avance",
        "Justificacion", "Observaciones",
    ]

    data = []
    task_n = 1
    for ao_code, ao_name, _, _ in AO_DATA[:3]:
        for cod, _, desc in CLASIFICADORES[:2]:
            pim = 50000
            prog = 45000
            ejec = 38000
            saldo = prog - ejec
            avance = round(ejec / prog * 100, 1)
            data.append((
                META_CODIGO, META_DESC, ao_code, ao_name,
                f"T{task_n:03d}", f"Actividad {task_n}", cod, desc,
                pim, prog, ejec, saldo, avance,
                "Ejecucion dentro de parametros normales", "Sin observaciones",
            ))
            task_n += 1

    write_headers_and_data(ws, headers, data, header_row=7)
    path = EJEMPLOS_DIR / "ejemplo_formato3.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} tareas)")


# =========================================================================
# 6. FORMATO 04 — Modificaciones Presupuestales
# =========================================================================
def generate_formato04():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 04"

    # Context matching formato04_parser positions
    ws["A1"] = "Formato 04 - Modificaciones Presupuestales"
    style_title(ws["A1"])
    ws.merge_cells("A1:F1")

    # Row 2: ue context
    ws.cell(row=2, column=1, value="Unidad Ejecutora:").font = LABEL_FONT
    ws.cell(row=2, column=3, value=UE_NOMBRE).font = DATA_FONT
    ws.cell(row=2, column=6, value=UE_CODIGO).font = DATA_FONT

    # Row 3: nota numero
    ws.cell(row=3, column=1, value="Nota Modificacion:").font = LABEL_FONT
    ws.cell(row=3, column=3, value="NM-2026-001").font = DATA_FONT
    ws.cell(row=3, column=6, value="15/01/2026").font = DATA_FONT

    # Row 4: anio
    ws.cell(row=4, column=1, value="Ano Fiscal:").font = LABEL_FONT
    ws.cell(row=4, column=6, value=ANIO).font = DATA_FONT

    headers = ["Clasificador", "Descripcion", "Asignado", "Habilitadora", "Habilitada", "PIM Resultante"]
    data = [
        ("2.3.1.5.1.1", "MATERIALES Y UTILES DE OFICINA", 160000, 12000, 0, 172000),
        ("2.3.2.7.1.1", "SERVICIOS DE CONSULTORIA", 530000, 0, 12000, 518000),
        ("2.3.2.2.1.1", "SERVICIO DE ENERGIA ELECTRICA", 240000, 5000, 0, 245000),
        ("2.3.2.1.2.1", "PASAJES Y GASTOS DE TRANSPORTE", 137000, 0, 5000, 132000),
        ("2.6.3.2.1.1", "EQUIPOS COMPUTACIONALES", 190000, 20000, 0, 210000),
        ("2.3.2.7.11.99", "OTROS SERVICIOS DIVERSOS", 84000, 0, 20000, 64000),
    ]

    write_headers_and_data(ws, headers, data, header_row=7)
    fmt_money(ws, 8, 7 + len(data), [3, 4, 5, 6])

    path = EJEMPLOS_DIR / "ejemplo_formato04.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} modificaciones)")


# =========================================================================
# 7. FORMATO 5A — Programacion AO (solo programado)
# =========================================================================
def generate_formato5a():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 5.A"

    # Context matching formato5a_parser positions:
    # (2, 2) = row 3 col C = ue_nombre, (2, 5) = row 3 col F = ue_codigo
    # (3, 5) = row 4 col F = meta_codigo, (4, 5) = row 5 col F = anio
    ws["A1"] = "Formato 5.A - Programacion de Actividades Operativas"
    style_title(ws["A1"])
    ws.merge_cells("A1:N1")

    ws.cell(row=3, column=1, value="Unidad Ejecutora:").font = LABEL_FONT
    ws.cell(row=3, column=3, value="INEI SEDE CENTRAL").font = DATA_FONT
    ws.cell(row=3, column=6, value=UE_CODIGO).font = DATA_FONT

    ws.cell(row=4, column=1, value="Meta Presupuestal:").font = LABEL_FONT
    ws.cell(row=4, column=6, value=META_CODIGO).font = DATA_FONT

    ws.cell(row=5, column=1, value="Ano Fiscal:").font = LABEL_FONT
    ws.cell(row=5, column=6, value=ANIO).font = DATA_FONT

    # Headers at row 11 (data starts at row 12)
    headers = [
        "Codigo AO", "Nombre AO",
        "Ene", "Feb", "Mar", "Abr", "May", "Jun",
        "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
        "Total Programado",
    ]

    data = []
    for ao_code, ao_name, _, _ in AO_DATA:
        monthly = AO_MONTHLY_PROG[ao_code]
        total = sum(monthly)
        data.append((ao_code, ao_name, *monthly, total))

    write_headers_and_data(ws, headers, data, header_row=11)
    fmt_money(ws, 12, 11 + len(data), list(range(3, 16)))

    path = EJEMPLOS_DIR / "ejemplo_formato5a.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} AOs)")


# =========================================================================
# 8. FORMATO 5B — Ejecucion AO (triple: programado/ejecutado/saldo x 12)
# =========================================================================
def generate_formato5b():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 5.B"

    # Context rows (same positions as 5A)
    ws["A1"] = "Formato 5.B - Ejecucion de Actividades Operativas"
    style_title(ws["A1"])
    ws.merge_cells("A1:Z1")

    ws.cell(row=3, column=1, value="Unidad Ejecutora:").font = LABEL_FONT
    ws.cell(row=3, column=3, value="INEI SEDE CENTRAL").font = DATA_FONT
    ws.cell(row=3, column=6, value=UE_CODIGO).font = DATA_FONT

    ws.cell(row=4, column=1, value="Meta Presupuestal:").font = LABEL_FONT
    ws.cell(row=4, column=6, value=META_CODIGO).font = DATA_FONT

    ws.cell(row=5, column=1, value="Ano Fiscal:").font = LABEL_FONT
    ws.cell(row=5, column=6, value=ANIO).font = DATA_FONT

    # Two-row compound header (rows 9-10 = 0-based 8-9)
    # Row 9: month names spanning 3 columns each
    months_es = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # Row 9: Codigo AO, Nombre AO, then month names
    r9 = 9
    ws.cell(row=r9, column=1, value="Codigo AO").font = HEADER_FONT
    ws.cell(row=r9, column=1).fill = HEADER_FILL
    ws.cell(row=r9, column=2, value="Nombre AO").font = HEADER_FONT
    ws.cell(row=r9, column=2).fill = HEADER_FILL

    col = 3
    for month in months_es:
        for _ in range(3):
            cell = ws.cell(row=r9, column=col, value=month)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            col += 1

    # Row 10: sub-headers (Programado, Ejecutado, Saldo) repeated
    r10 = 10
    ws.cell(row=r10, column=1, value="").fill = HEADER_FILL
    ws.cell(row=r10, column=2, value="").fill = HEADER_FILL

    col = 3
    for _ in range(12):
        for sub in ["Programado", "Ejecutado", "Saldo"]:
            cell = ws.cell(row=r10, column=col, value=sub)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            col += 1

    # Data starts at row 12 (row 11 is separator)
    row_start = 12
    for i, (ao_code, ao_name, _, _) in enumerate(AO_DATA):
        r = row_start + i
        ws.cell(row=r, column=1, value=ao_code).font = DATA_FONT
        ws.cell(row=r, column=2, value=ao_name).font = DATA_FONT

        monthly = AO_MONTHLY_PROG[ao_code]
        col = 3
        for m_idx in range(12):
            prog = monthly[m_idx]
            ejec = round(prog * EXEC_RATIOS[m_idx])
            saldo = prog - ejec

            ws.cell(row=r, column=col, value=prog).font = DATA_FONT
            ws.cell(row=r, column=col + 1, value=ejec).font = DATA_FONT
            ws.cell(row=r, column=col + 2, value=saldo).font = DATA_FONT
            col += 3

    # Set column widths
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 35
    for c in range(3, 3 + 36):
        ws.column_dimensions[get_column_letter(c)].width = 12

    path = EJEMPLOS_DIR / "ejemplo_formato5b.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(AO_DATA)} AOs x 12 meses)")


# =========================================================================
# 9. FORMATO 5 RESUMEN
# =========================================================================
def generate_formato5_resumen():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 5 Resumen"

    write_context(ws, "Formato 5 Resumen - Ejecucion Consolidada AO", [
        ("Unidad Ejecutora:", UE_NOMBRE),
        ("Meta Presupuestal:", META_CODIGO),
        ("Ano Fiscal:", ANIO),
    ], 20)

    headers = [
        "Codigo AO", "Nombre AO", "PIM", "CCP",
        "Compromiso Anual", "Devengado", "Girado",
        "Saldo", "% Avance PIM", "Semaforo",
    ]

    data = []
    for ao_code, ao_name, _, _ in AO_DATA:
        monthly = AO_MONTHLY_PROG[ao_code]
        pim = sum(monthly)
        # Only Jan+Feb executed
        ejec_total = round(monthly[0] * 0.95 + monthly[1] * 0.85)
        ccp = round(ejec_total * 1.15)
        compromiso = round(ejec_total * 1.05)
        devengado = ejec_total
        girado = round(ejec_total * 0.95)
        saldo = pim - devengado
        avance = round(devengado / pim * 100, 1)
        semaforo = "VERDE" if avance >= 15 else "AMARILLO" if avance >= 10 else "ROJO"
        data.append((ao_code, ao_name, pim, ccp, compromiso, devengado, girado, saldo, avance, semaforo))

    write_headers_and_data(ws, headers, data, header_row=6)
    path = EJEMPLOS_DIR / "ejemplo_formato5_resumen.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} AOs)")


# =========================================================================
# 10. ANEXO 01 — Recursos Humanos
# =========================================================================
def generate_anexo01():
    wb = Workbook()
    ws = wb.active
    ws.title = "Anexo 01"

    write_context(ws, "Anexo 01 - Datos de Recursos Humanos", [
        ("Unidad Ejecutora:", UE_NOMBRE),
        ("Meta Presupuestal:", META_CODIGO),
        ("Ano Fiscal:", ANIO),
    ], 12)

    headers = [
        "N", "DNI", "Apellidos y Nombres", "Cargo",
        "Area", "Regimen Laboral", "Tipo Contrato",
        "Fecha Inicio", "Fecha Fin", "Remuneracion Mensual",
        "Observaciones", "Estado",
    ]

    data = [
        (1, "45678901", "GARCIA LOPEZ, MARIA ELENA", "Especialista TI",
         "OTIN", "728", "CAS", "2025-01-15", "2026-12-31", 5500.00, "", "ACTIVO"),
        (2, "34567890", "TORRES MARTINEZ, CARLOS", "Analista Programador",
         "OTIN", "728", "CAS", "2025-03-01", "2026-12-31", 4800.00, "", "ACTIVO"),
        (3, "23456789", "QUISPE HUAMAN, ANA ROSA", "Asistente Administrativo",
         "OTA", "276", "NOMBRADO", "2018-04-01", "", 3200.00, "", "ACTIVO"),
        (4, "12345678", "MENDOZA CHAVEZ, JUAN PEDRO", "Director Tecnico",
         "DEC", "728", "CAS", "2024-07-01", "2026-06-30", 8500.00, "", "ACTIVO"),
        (5, "56789012", "VARGAS ROJAS, PATRICIA", "Especialista Presupuesto",
         "OTPP", "276", "NOMBRADO", "2015-01-15", "", 4200.00, "", "ACTIVO"),
        (6, "67890123", "HUAMANI CONDORI, LUIS", "Tecnico Estadistico",
         "DEC", "728", "CAS", "2025-06-01", "2026-12-31", 3800.00, "", "ACTIVO"),
        (7, "78901234", "SILVA RAMOS, CARMEN", "Secretaria Ejecutiva",
         "OTIN", "276", "NOMBRADO", "2010-03-01", "", 3000.00, "", "ACTIVO"),
        (8, "89012345", "RIVERA SANTOS, MIGUEL", "Operador de Equipos",
         "OTIN", "728", "CAS", "2025-09-01", "2026-12-31", 2800.00, "Medio tiempo", "ACTIVO"),
    ]

    write_headers_and_data(ws, headers, data, header_row=7)
    path = EJEMPLOS_DIR / "ejemplo_anexo01.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} registros RRHH)")


# =========================================================================
# 11. SIAF (Ejecucion Presupuestal)
# =========================================================================
def generate_siaf():
    wb = Workbook()
    ws = wb.active
    ws.title = "SIAF"

    # SIAF parser: data_start_row=4, scans rows 0-3 for keywords
    # Headers must be at row 4 (index 3) or earlier
    # Row 1: title (no conflicting keywords)
    # Row 2: context (entidad + UE)
    # Row 3: context (ejercicio)
    # Row 4: HEADERS ← must be here for parser to find them
    ws["A1"] = "SIAF - Reporte"
    style_title(ws["A1"])
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 28
    ws.cell(row=2, column=1, value="Entidad:").font = LABEL_FONT
    ws.cell(row=2, column=2, value="001-INEI").font = DATA_FONT
    ws.cell(row=2, column=5, value="UE:").font = LABEL_FONT
    ws.cell(row=2, column=6, value=UE_NOMBRE).font = DATA_FONT
    ws.cell(row=3, column=1, value="Ejercicio:").font = LABEL_FONT
    ws.cell(row=3, column=2, value=ANIO).font = DATA_FONT

    headers = [
        "Anio", "Clasificador", "Descripcion",
        "PIA", "PIM", "Certificado",
        "Compromiso Anual", "Devengado", "Girado",
    ]

    # SIAF provides execution data for the SAME clasificadores as Formato 1
    # This is the key — it updates the zeros left by Formato 1
    data = []
    for cod, _, desc in CLASIFICADORES:
        monthly = MONTHLY_BUDGETS.get(cod, [0]*12)
        pim = sum(monthly)
        pia = int(pim * 0.95)

        # Execution = Jan fully + Feb ~85% (we're in Feb 2026)
        jan = monthly[0]
        feb = monthly[1]
        devengado = round(jan * 0.95 + feb * 0.82)
        certificado = round(devengado * 1.12)
        compromiso = round(devengado * 1.05)
        girado = round(devengado * 0.96)

        data.append((ANIO, cod, desc, pia, pim, certificado, compromiso, devengado, girado))

    write_headers_and_data(ws, headers, data, header_row=4)
    fmt_money(ws, 5, 4 + len(data), list(range(4, 10)))

    path = EJEMPLOS_DIR / "ejemplo_siaf.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} registros ejecucion)")


# =========================================================================
# 12. SIGA (Requerimientos Logisticos)
# =========================================================================
def generate_siga():
    wb = Workbook()
    ws = wb.active
    ws.title = "SIGA"

    write_context(ws, "SIGA - Requerimientos Logisticos", [
        ("Entidad:", "001-INEI"),
        ("Periodo:", "Enero - Diciembre 2026"),
    ], 9)

    headers = [
        "Nro. Requerimiento", "Descripcion", "Unidad Medida",
        "Cantidad", "Precio Unitario", "Monto Total",
        "Estado", "Proveedor", "Fecha",
    ]
    data = [
        ("REQ-2026-0001", "Toner HP LaserJet Pro M404dn", "UNIDAD",
         50, 280.00, 14000.00, "ATENDIDO", "SOLUTEK SAC", "2026-01-15"),
        ("REQ-2026-0002", "Papel Bond A4 75gr (millar)", "MILLAR",
         200, 22.50, 4500.00, "ATENDIDO", "TAI LOY SA", "2026-01-20"),
        ("REQ-2026-0003", "Laptop Dell Latitude 5540", "UNIDAD",
         15, 4800.00, 72000.00, "EN PROCESO", "DELL PERU SRL", "2026-02-01"),
        ("REQ-2026-0004", "Servicio de limpieza mensual", "SERVICIO",
         12, 8500.00, 102000.00, "ATENDIDO", "LIMPIEZA TOTAL SAC", "2026-01-05"),
        ("REQ-2026-0005", "Archivadores de palanca oficio", "UNIDAD",
         500, 6.50, 3250.00, "ATENDIDO", "TAI LOY SA", "2026-02-10"),
        ("REQ-2026-0006", "Monitor LG 24'' IPS Full HD", "UNIDAD",
         20, 650.00, 13000.00, "PENDIENTE", "LG ELECTRONICS PERU", "2026-02-15"),
        ("REQ-2026-0007", "Servicio de mantenimiento ascensores", "SERVICIO",
         4, 3200.00, 12800.00, "EN PROCESO", "SCHINDLER DEL PERU", "2026-03-01"),
        ("REQ-2026-0008", "Memoria RAM DDR4 16GB", "UNIDAD",
         30, 185.00, 5550.00, "PENDIENTE", "SOLUTEK SAC", "2026-03-05"),
        ("REQ-2026-0009", "Sillas ergonomicas de oficina", "UNIDAD",
         25, 890.00, 22250.00, "EN PROCESO", "OFIMUEBLES PERU SAC", "2026-02-20"),
        ("REQ-2026-0010", "Servicio internet fibra optica 500Mbps", "SERVICIO",
         12, 2500.00, 30000.00, "ATENDIDO", "MOVISTAR PERU", "2026-01-10"),
    ]

    write_headers_and_data(ws, headers, data, header_row=4)
    fmt_money(ws, 5, 4 + len(data), [5, 6])

    path = EJEMPLOS_DIR / "ejemplo_siga.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path} ({len(data)} requerimientos)")


# =========================================================================
# Plantillas SIAF / SIGA (empty templates)
# =========================================================================
def generate_plantilla_siaf():
    wb = Workbook()
    ws = wb.active
    ws.title = "SIAF"
    ws["A1"] = "SIAF - Reporte"
    style_title(ws["A1"])
    ws.merge_cells("A1:I1")
    ws.cell(row=2, column=1, value="Entidad:").font = LABEL_FONT
    ws.cell(row=2, column=2, value="").font = DATA_FONT
    ws.cell(row=3, column=1, value="Ejercicio:").font = LABEL_FONT
    ws.cell(row=3, column=2, value=2026).font = DATA_FONT
    headers = ["Anio", "Clasificador", "Descripcion", "PIA", "PIM",
               "Certificado", "Compromiso Anual", "Devengado", "Girado"]
    from openpyxl.utils import get_column_letter
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(h) + 4)
    path = PLANTILLAS_DIR / "plantilla_siaf.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path}")


def generate_plantilla_siga():
    wb = Workbook()
    ws = wb.active
    ws.title = "SIGA"
    write_context(ws, "SIGA - Requerimientos Logisticos", [
        ("Entidad:", ""), ("Periodo:", ""),
    ], 9)
    headers = ["Nro. Requerimiento", "Descripcion", "Unidad Medida",
               "Cantidad", "Precio Unitario", "Monto Total",
               "Estado", "Proveedor", "Fecha"]
    from openpyxl.utils import get_column_letter
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(h) + 4)
    path = PLANTILLAS_DIR / "plantilla_siga.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"  [OK] {path}")


# =========================================================================
# Main
# =========================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("Generando TODOS los archivos ejemplo (12 formatos)")
    print("=" * 60)
    print()

    print("[1. Datos Maestros]")
    generate_cuadro_ao_meta()
    generate_tablas()
    print()

    print("[2. Formatos DDNNTT]")
    generate_formato1()
    generate_formato2()
    generate_formato3()
    generate_formato04()
    generate_formato5a()
    generate_formato5b()
    generate_formato5_resumen()
    generate_anexo01()
    print()

    print("[3. Sistemas Externos]")
    generate_siaf()
    generate_siga()
    print()

    print("[4. Plantillas SIAF/SIGA]")
    generate_plantilla_siaf()
    generate_plantilla_siga()
    print()

    print("=" * 60)
    print("LISTO! 12 archivos ejemplo + 2 plantillas generados.")
    print(f"  Ejemplos: {EJEMPLOS_DIR}")
    print(f"  Plantillas: {PLANTILLAS_DIR}")
    print()
    print("Orden de carga recomendado:")
    print("  1. ejemplo_cuadro_ao_meta.xlsx (datos maestros)")
    print("  2. ejemplo_tablas.xlsx (clasificadores)")
    print("  3. ejemplo_formato1.xlsx (programacion anual)")
    print("  4. ejemplo_formato5a.xlsx (programacion AO)")
    print("  5. ejemplo_formato5b.xlsx (ejecucion AO)")
    print("  6. ejemplo_siaf.xlsx (ejecucion presupuestal)")
    print("  7. Los demas formatos en cualquier orden")
    print("=" * 60)
